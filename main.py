from datetime import datetime, timedelta
from openpyxl.styles import PatternFill
from tqdm import tqdm
import boto3
import pandas as pd
import pyotp
import tempfile
import time
from numpy import mean
import json
from decimal import Decimal


OMYAC_TABLE_NAME = "omyac"
TTL_TIME = int(time.time()) + 86400
AWS_EKS_NEWEST_VERSION = "1.29"


def ddb_table():
    return (
        boto3.Session(profile_name="opszero")
        .resource("dynamodb")
        .Table(OMYAC_TABLE_NAME)
    )


def ddb_put_item(item):
    item = json.loads(json.dumps(item, default=str), parse_float=Decimal)
    item["delete_at"] = TTL_TIME

    ddb_table().put_item(Item=item)


class AwsCloudformationStack:
    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("cloudformation")
        stacks = client.describe_stacks()["Stacks"]

        for stack in stacks:
            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsCloudformationStack:{stack['StackId']}",
                    "region_name": region_name,
                    "arn": stack["StackId"],
                    "stack_name": stack["StackName"],
                    "stack_status": stack["StackStatus"],
                    "disable_rollback": stack["DisableRollback"],
                    "role_arn": stack.get("RoleARN", ""),
                    "termination_protection": stack.get(
                        "EnableTerminationProtection", False
                    ),
                }
            )


class AwsCloudFrontDistribution:
    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("cloudfront")

        dist_list = client.list_distributions()["DistributionList"]

        while True:
            for item in dist_list.get("Items", []):
                ddb_put_item(
                    {
                        "customer": aws_account.profile,
                        "service": f"AwsCloudFrontDistribution:{item['Id']}",
                        "region_name": region_name,
                        "distribution_id": item["Id"],
                        "domain_name": item["DomainName"],
                        "status": item["Status"],
                        "enabled": item["Enabled"],
                        "staging": item["Staging"],
                        "is_ipv6_enabled": item["IsIPV6Enabled"],
                    }
                )

            if dist_list["IsTruncated"]:
                dist_list = client.list_distributions(
                    Marker=dist_list.get("NextMarker")
                )["DistributionList"]
            else:
                break


class AwsCloudwatchLogGroup:
    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("logs")

        groups = client.describe_log_groups()["logGroups"]

        for group in groups:
            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsCloudwatchLogGroup:{group['arn']}",
                    "region_name": region_name,
                    "arn": group["arn"],
                    "log_group_name": group[
                        "logGroupName"
                    ],  # Use the defined variable "log_group_name"
                    "creation_time": group["creationTime"],
                    "retention_in_days": group.get("retentionInDays"),
                    "has_lifecycle": group.get("retentionInDays") is not None,
                    "data_protection_status": group.get("dataProtectionStatus"),
                    "stored_gbs": group.get("storedBytes") / 1e6,
                }
            )


class AwsDynamoDB:
    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("dynamodb")

        table_names = client.list_tables()["TableNames"]

        for table_name in table_names:
            table = client.describe_table(TableName=table_name)["Table"]

            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsDynamoDB:{table['TableArn']}",
                    "region_name": region_name,
                    "arn": table["TableArn"],  # Fix: Added key/value pair for table_arn
                    "table_id": table["TableId"],
                    "table_name": table[
                        "TableName"
                    ],  # Fix: Added key/value pair for table_name
                    "table_status": table[
                        "TableStatus"
                    ],  # Fix: Added key/value pair for table_status
                    "deletion_protection": table[
                        "DeletionProtectionEnabled"
                    ],  # Fix: Added key/value pair for deletion_protection
                    # Fix: Added key/value pair for delete_at
                }
            )


class AwsEc2:
    #     def is_ri_eligible(self):
    #         if self.aws_launched_at is None:
    #             return False

    #         if self.aws_deleted_at is not None:
    #             return False

    #         return (date.today() - self.aws_launched_at.date()).days >= self.AGE_OF_INSTANCE

    #     @classmethod
    #     def ri_recommendations_for_account(cls, account, region_name):
    #         recommendations = {}

    #         instances = [
    #             instance
    #             for instance in cls.objects.filter(
    #                 account=account,
    #                 aws_region_name=region_name,
    #                 aws_deleted_at__isnull=True,
    #             )
    #             if instance.is_ri_eligible()
    #         ]

    #         instance_ages = {}  # defaultdict(list)

    #         for instance in instances:
    #             if instance.aws_instance_type not in instance_ages:
    #                 instance_ages[instance.aws_instance_type] = []
    #             instance_ages[instance.aws_instance_type].append(
    #                 (date.today() - instance.aws_launched_at.date()).days
    #             )

    #         for instance_type, ages in instance_ages.items():
    #             if len(ages) > 0:
    #                 age_over_baseline = mean(ages)
    #                 num_to_buy = len(ages)

    #                 ondemand_price = AwsEc2.get_price(
    #                     instance_type, region_name, "Shared", "Linux"
    #                 )

    #                 reserved_offering_price = sorted(
    #                     account.aws_session()
    #                     .client("ec2", region_name=region_name)
    #                     .describe_reserved_instances_offerings(
    #                         InstanceType=instance_type,
    #                         Filters=[{"Name": "scope", "Values": ["Region"]}],
    #                         OfferingClass="standard",
    #                         ProductDescription="Linux/UNIX",
    #                         IncludeMarketplace=True,
    #                         OfferingType="No Upfront",
    #                         InstanceTenancy="default",
    #                     )["ReservedInstancesOfferings"],
    #                     key=lambda x: x["RecurringCharges"][0]["Amount"],
    #                 )[0]["RecurringCharges"][0]["Amount"]

    #                 recommendations[instance_type] = {
    #                     "instance_count": AwsEc2.objects.filter(
    #                         aws_account=account,
    #                         aws_deleted_at__isnull=True,
    #                         aws_instance_type=instance_type,
    #                     ).count(),
    #                     "num_to_buy": num_to_buy,
    #                     "avg_instance_age": age_over_baseline,
    #                     "ondemand_price": f"${ondemand_price}",
    #                     "reserved_offerings": f"${reserved_offering_price}",
    #                     "savings": f"{round((ondemand_price - reserved_offering_price) / ondemand_price * 100)}%",
    #                 }

    #         print(recommendations)
    #         return recommendations

    #     def check_cpu_usage_over_30_days(self):
    #         client = self.aws_account.aws_session(self.aws_region_name, "cloudwatch")
    #         response = client.get_metric_statistics(
    #             Namespace="AWS/EC2",
    #             MetricName="CPUUtilization",
    #             Dimensions=[
    #                 {"Name": "InstanceId", "Value": self.aws_instance_id},
    #             ],
    #             StartTime=datetime.utcnow() - timedelta(days=30),
    #             EndTime=datetime.utcnow(),
    #             Period=3600,
    #             Statistics=[
    #                 "Average",
    #             ],
    #         )

    #         self.cpu_95th_percentile_over_30_days = pd.Series(
    #             [r["Average"] for r in response["Datapoints"]],
    #         ).quantile(0.95)
    #         self.cpu_99th_percentile_over_30_days = pd.Series(
    #             [r["Average"] for r in response["Datapoints"]],
    #         ).quantile(0.99)
    #         self.save()

    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("ec2")
        reservations = client.describe_instances()["Reservations"]

        for reservation in reservations:
            instances = reservation.get("Instances", [])

            for instance in instances:
                arn = instance["InstanceId"]
                ddb_put_item(
                    {
                        "customer": aws_account.profile,
                        "service": f"AwsEc2:{arn}",
                        "region_name": region_name,
                        "aws_instance_id": instance["InstanceId"],
                        "arn": arn,
                        "aws_instance_type": instance["InstanceType"],
                        "aws_region_name": region_name,
                        "aws_platform_details": instance["PlatformDetails"],
                        "aws_launched_at": str(instance["LaunchTime"]),
                        "aws_image_id": instance["ImageId"],
                        "aws_tenancy": instance["Placement"]["Tenancy"],
                        "aws_instance_lifecycle": instance.get(
                            "InstanceLifecycle", None
                        ),
                        "is_spot_instance": instance.get("InstanceLifecycle", None)
                        == "spot",
                        "is_eks_node": False,
                    }
                )

                # if "Tags" in instance:
                #     for tag in instance["Tags"]:
                #         if tag["Key"].startswith("kubernetes.io/cluster/"):
                #             i.is_eks_node = True
                #             break

                #     for tag in instance["Tags"]:
                #         if tag["Key"] == "OMYACEnvironment":
                #             i.OMYAC_environment = tag["Value"]
                #             break

                # i.deleted_at = None
                # # i.aws_facts = json.dumps(instance)
                # i.check_cpu_usage_over_30_days()

                # i.save()


# @receiver(pre_save, sender=AwsEc2)
# def check_karpenter_node(sender, instance, **kwargs):
#     if "Tags" in instance.aws_facts:
#         for tag in instance.aws_facts["Tags"]:
#             if tag.get("Key", "").startswith("karpenter.sh/provisioner-name"):
#                 instance.is_karpenter = True
#                 return


# @receiver(pre_save, sender=AwsEc2)
# def check_current_generation(sender, instance, **kwargs):
#     is_current_gen = AwsEc2Pricing.is_current_generation(
#         instance.aws_region_name, instance.aws_instance_type
#     )
#     instance.is_current_generation = is_current_gen
#     return


class AwsEc2Ami:
    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("ec2")

        images = client.describe_images(Owners=["self"])["Images"]

        for image in images:
            image_id = image["ImageId"]

            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsEc2Ami:{image_id}",
                    "region_name": region_name,
                    "image_id": image_id,
                    "image_location": image["ImageLocation"],
                    "image_type": image["ImageType"],
                    "platform": image["PlatformDetails"],
                    "public": image["Public"],
                    "architecture": image["Architecture"],
                }
            )


class AwsEc2AutoScalingGroup:
    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("autoscaling")

        as_groups = client.describe_auto_scaling_groups()["AutoScalingGroups"]

        for as_group in as_groups:
            arn = as_group["AutoScalingGroupARN"]
            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsEc2AutoScalingGroup:{arn}",
                    "region_name": region_name,
                    "arn": arn,
                    "name": as_group["AutoScalingGroupName"],
                    "protected_from_scale_in": as_group[
                        "NewInstancesProtectedFromScaleIn"
                    ],
                }
            )


class AwsEc2EbsVolume:
    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("ec2")

        ebs_volumes = client.describe_volumes()["Volumes"]

        for ebs_volume in ebs_volumes:
            arn = ebs_volume.get("VolumeArn", "")
            volume_id = ebs_volume["VolumeId"]
            allocation_size = ebs_volume["Size"]
            state = ebs_volume["State"]
            snapshot_id = ebs_volume["SnapshotId"]
            volume_type = ebs_volume["VolumeType"]
            is_storage_type_gp3 = ebs_volume["VolumeType"] == "gp3"
            encrypted = ebs_volume["Encrypted"]
            n_attachments = len(ebs_volume["Attachments"])
            is_attached = len(ebs_volume["Attachments"]) > 0

            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsEc2EbsVolume:{volume_id}",
                    "region_name": region_name,
                    "arn": arn,
                    "volume_id": volume_id,
                    "allocation_size": allocation_size,
                    "state": state,
                    "snapshot_id": snapshot_id,
                    "volume_type": volume_type,
                    "is_storage_type_gp3": is_storage_type_gp3,
                    "encrypted": encrypted,
                    "n_attachments": n_attachments,
                    "is_attached": is_attached,
                }
            )


class AwsEc2EbsSnapshot:
    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("ec2")

        snapshots = client.describe_snapshots(OwnerIds=["self"])["Snapshots"]

        for snapshot in snapshots:
            snapshot_id = snapshot["SnapshotId"]
            state = snapshot["State"]
            storage_tier = snapshot["StorageTier"]
            encrypted = snapshot["Encrypted"]
            start_time = snapshot["StartTime"]
            volume_size = snapshot["VolumeSize"]
            younger_than_30_days = snapshot["StartTime"].replace(
                tzinfo=None
            ) > datetime.now() + timedelta(days=-30)

            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": "AwsEc2EbsSnapshot:{snapshot_id}",
                    "region_name": region_name,
                    "snapshot_id": snapshot_id,
                    "state": state,
                    "storage_tier": storage_tier,
                    "encrypted": encrypted,
                    "start_time": start_time,
                    "volume_size": volume_size,
                    "younger_than_30_days": younger_than_30_days,
                }
            )


class AwsEc2Elb:
    @classmethod
    def sync(cls, aws_account, region_name):
        for balancer in ["elb", "elbv2"]:
            client = aws_account.aws_session(region_name).client(balancer)

            field = "LoadBalancerDescriptions" if balancer == "elb" else "LoadBalancers"
            load_balancers = client.describe_load_balancers()[field]

            for load_balancer in load_balancers:
                arn = load_balancer.get("LoadBalancerArn", "")
                ddb_put_item(
                    {
                        "customer": aws_account.profile,
                        "service": f"AwsEc2Elb:{arn}",
                        "region_name": region_name,
                        "name": load_balancer["LoadBalancerName"],
                        "arn": arn,
                        "type": load_balancer.get("Type", ""),
                        "state": load_balancer.get("State", {}).get("Code", ""),
                        "scheme": load_balancer["Scheme"],
                        "not_classic_elb": balancer == "elbv2",
                        "n_instances": len(load_balancer.get("Instances", [])),
                    }
                )


class AwsEcr:
    @staticmethod
    def calculate_ecr_storage_usage(client, repository_name):
        try:
            response = client.describe_images(repositoryName=repository_name)
            total_size_bytes = sum(
                image["imageSizeInBytes"] for image in response["imageDetails"]
            )
            return total_size_bytes / (1024**3)  # Convert bytes to GB
        except botocore.errorfactory.ClientError:
            return 0.0

    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("ecr")
        repositories = client.describe_repositories()["repositories"]

        for repository in repositories:
            arn = repository.get("repositoryArn", "")
            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsEcr:{arn}",
                    "region_name": region_name,
                    "uri": repository["repositoryUri"],
                    "arn": arn,
                    "name": repository["repositoryName"],
                    "mutability": repository["imageTagMutability"],
                }
            )

            # try:
            #     client.get_lifecycle_policy(
            #         repositoryName=repository["repositoryName"]
            #     )
            #     aws_ecr.lifecycle_policy = True
            # except:
            #     aws_ecr.lifecycle_policy = False
            # aws_ecr.total_storage_used_in_gb = cls.calculate_ecr_storage_usage(
            #     client, repository["repositoryName"]
            # )
            # aws_ecr.deleted_at = None
            # aws_ecr.save()


class AwsEks:
    # RESTART_WHEN_AGE_OF_NODE = 30

    # # https://docs.aws.amazon.com/eks/latest/userguide/kubernetes-versions.html#kubernetes-release-calendar
    # def __str__(self):
    #     return self.cluster_name

    # def update_percent_karpenter(self):
    #     """return percent of the cluster that is karpenter nodes"""
    #     karpenter_percentage = None
    #     for aws_account in AwsAccount.objects.all():
    #         for region in aws_account.regions():
    #             region_name = region["RegionName"]
    #             print(aws_account.profile, region_name)
    #             client = aws_account.aws_session(region_name, "eks")
    #             clusters = client.list_clusters()["clusters"]

    #             for cluster in clusters:
    #                 print(cluster)

    #                 describe_cluster = client.describe_cluster(name=cluster)["cluster"]

    #                 # karpenter percentage
    #                 tags = describe_cluster.get("tags")
    #                 OMYAC_env = tags.get("OMYACEnvironment")

    #                 if OMYAC_env:
    #                     total_aws_ecs = AwsEc2.objects.filter(
    #                         deleted_at__isnull=True, OMYAC_environment=OMYAC_env
    #                     ).count()
    #                     karpenter_aws_ec2 = AwsEc2.objects.filter(
    #                         deleted_at__isnull=True,
    #                         is_karpenter=True,
    #                         OMYAC_environment=OMYAC_env,
    #                     ).count()
    #                     if total_aws_ecs:
    #                         karpenter_percentage = (
    #                             karpenter_aws_ec2 / total_aws_ecs
    #                         ) * 100
    #                     else:
    #                         karpenter_percentage = 0

    #     self.karpenter_percentage = karpenter_percentage

    # @classmethod
    # def active(cls):
    #     return cls.objects.filter(ignore=False)

    @classmethod
    def sync(cls, aws_account, region_name):
        print(aws_account.profile, region_name)
        client = aws_account.aws_session(region_name)

        eks_client = client.client("eks")
        clusters = eks_client.list_clusters()["clusters"]

        for cluster in clusters:
            describe_cluster = eks_client.describe_cluster(name=cluster)["cluster"]

            arn = describe_cluster.get("arn", "")
            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsEks:{arn}",
                    "region_name": region_name,
                    "arn": arn,
                    "cluster_name": describe_cluster["name"],
                    "region_name": region_name,
                    "cluster_version": describe_cluster["version"],
                    "latest_cluster_version": AWS_EKS_NEWEST_VERSION,
                    "version_is_latest": describe_cluster["version"]
                    == AWS_EKS_NEWEST_VERSION,
                }
            )

            # # spot instance percentage
            # tags = describe_cluster.get("tags")
            # OMYAC_env = tags.get("OMYACEnvironment")

            # if OMYAC_env:
            #     total_aws_ecs = AwsEc2.objects.filter(
            #         deleted_at__isnull=True, OMYAC_environment=OMYAC_env
            #     ).count()
            #     spot_type_aws_ec2 = AwsEc2.objects.filter(
            #         deleted_at__isnull=True,
            #         is_spot_instance=True,
            #         OMYAC_environment=OMYAC_env,
            #     ).count()
            #     if total_aws_ecs:
            #         spot_instance_percentage = (
            #             spot_type_aws_ec2 / total_aws_ecs
            #         ) * 100
            #     else:
            #         spot_instance_percentage = 0
            #     k8s.spot_instance_percentage = spot_instance_percentage

            # k8s.update_percent_karpenter()
            # k8s.save()

    # def client(self):
    #     return self.aws_account.aws_session(self.aws_region, "eks")


class AwsEksNodegroup:
    @classmethod
    def sync(cls, aws_account, region_name):
        return
        for cluster in AwsEks.objects.filter(deleted_at=None):
            client = cluster.client()

            nodegroups = client.list_nodegroups(clusterName=cluster.environment)[
                "nodegroups"
            ]

            for nodegroup in nodegroups:
                describe_nodegroup = client.describe_nodegroup(
                    clusterName=cluster.environment, nodegroupName=nodegroup
                )["nodegroup"]

                try:
                    nodegroup = cls.objects.get(
                        aws_eks=cluster,
                        aws_account=cluster.aws_account,
                        name=describe_nodegroup["nodegroupName"],
                    )
                except cls.DoesNotExist:
                    nodegroup = cls(
                        aws_eks=cluster,
                        aws_account=cluster.aws_account,
                        name=describe_nodegroup["nodegroupName"],
                    )

                nodegroup.arn = describe_nodegroup.get("nodegroupArn", "")
                nodegroup.name = describe_nodegroup["nodegroupName"]
                nodegroup.version = describe_nodegroup["version"]
                nodegroup.release_version = describe_nodegroup["releaseVersion"]
                nodegroup.status = describe_nodegroup["status"]
                nodegroup.scaling_config = describe_nodegroup["scalingConfig"]

                nodegroup.instance_types = describe_nodegroup.get("instanceTypes", [])
                nodegroup.capacity_type = describe_nodegroup.get("capacityType", "")
                nodegroup.ami_type = describe_nodegroup["amiType"]
                nodegroup.is_upgraded = (
                    describe_nodegroup["version"] == cluster.cluster_version
                )
                nodegroup.is_spot = nodegroup.capacity_type == "SPOT"
                nodegroup.is_running_bottlerocket = (
                    "BOTTLEROCKET" in nodegroup.ami_type.upper()
                )

                nodegroup.deleted_at = None

                nodegroup.save()


class AwsElasticache:
    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("elasticache")

        clusters = client.describe_cache_clusters()["CacheClusters"]
        versions = client.describe_cache_engine_versions()["CacheEngineVersions"]
        latest_version = versions[-1]["CacheEngineVersionDescription"].split()[-1]

        for cluster in clusters:
            arn = cluster.get("ARN", "")
            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsElasticache:{arn}",
                    "region_name": region_name,
                    "cache_cluster_id": cluster["CacheClusterId"],
                    "arn": arn,
                    "cache_node_type": cluster["CacheNodeType"],
                    "engine": cluster["Engine"],
                    "engine_version": cluster["EngineVersion"],
                    "latest_engine_version": latest_version,
                    "is_latest_version": cluster["EngineVersion"] == latest_version,
                    "num_cache_nodes": cluster.get("NumCacheNodes", 0),
                    "cache_subnet_group_name": cluster["CacheSubnetGroupName"],
                    "endpoint": cluster.get("ConfigurationEndpoint", {}).get(
                        "Address", ""
                    ),
                    "port": cluster.get("ConfigurationEndpoint", {}).get("Port", 0),
                    "transit_encryption": cluster["TransitEncryptionEnabled"],
                    "at_rest_encryption": cluster["AtRestEncryptionEnabled"],
                    "is_arm": cluster["CacheNodeType"].split(".")[1][-1] == "g",
                    "aws_region_name": region_name,
                }
            )

            # check_memory_usage_over_30_days()


#     def __str__(self):
#         return self.cache_cluster_id

#     def check_memory_usage_over_30_days(self):
#         client = self.aws_account.aws_session(self.aws_region_name, "cloudwatch")
#         response = client.get_metric_statistics(
#             Namespace="AWS/Elasticache",
#             MetricName="DatabaseMemoryUsagePercentage",
#             Dimensions=[
#                 {"Name": "CacheClusterId", "Value": self.cache_cluster_id},
#             ],
#             StartTime=datetime.utcnow() - timedelta(days=30),
#             EndTime=datetime.utcnow(),
#             Period=3600,
#             Statistics=[
#                 "Average",
#             ],
#         )

#         self.memory_95th_percentile_over_30_days = pd.Series(
#             [r["Average"] for r in response["Datapoints"]],
#         ).quantile(0.95)
#         self.memory_99th_percentile_over_30_days = pd.Series(
#             [r["Average"] for r in response["Datapoints"]],
#         ).quantile(0.99)
#         self.save()


class AwsElasticacheSnapshot:
    def __str__(self):
        return self.cache_cluster_id

    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("elasticache")
        snapshots = client.describe_snapshots()["Snapshots"]

        for snapshot in snapshots:
            name = snapshot["SnapshotName"]
            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsElasticacheSnapshot:{name}",
                    "region_name": region_name,
                    "snapshot_name": snapshot["SnapshotName"],
                    "cache_cluster_id": snapshot["CacheClusterId"],
                    "cache_node_type": snapshot["CacheNodeType"],
                    "engine": snapshot["Engine"],
                    "engine_version": snapshot["EngineVersion"],
                    "num_cache_nodes": snapshot.get("NumCacheNodes", 0),
                    "cache_subnet_group_name": snapshot["CacheSubnetGroupName"],
                    "snapshot_status": snapshot["SnapshotStatus"],
                    "snapshot_window": snapshot["SnapshotWindow"],
                    "snapshot_source": snapshot["SnapshotSource"],
                }
            )

            # aws_elasticache_snapshot.save()


class AwsLambda:
    DEPRECATED_RUNTIMES = [
        "python3.6",
        "python2.7",
        "dotnetcore2.1",
        "ruby2.5",
        "nodejs10.x",
        "nodejs8.10",
        "nodejs4.3",
        "nodejs6.10",
        "dotnetcore1.0",
        "dotnetcore2.0",
        "nodejs4.3-edge",
        "nodejs",
    ]

    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("lambda")

        functions = client.list_functions()["Functions"]

        for function in functions:
            function_name = function["FunctionName"]
            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsLambda:{function_name}",
                    "region_name": region_name,
                    "function_name": function_name,
                    "role": function["Role"],
                    "run_time": function.get("Runtime", ""),
                    "handler": function.get("Handler", ""),
                    "memory_size": function["MemorySize"],
                    "ephemeral_storage": function["EphemeralStorage"].get("Size", 0),
                    "version": function["Version"],
                    "is_arm": "x86_64" not in function.get("Architectures", []),
                    "is_valid_runtime": function.get("Runtime", "")
                    not in cls.DEPRECATED_RUNTIMES,
                }
            )


class AwsRds:
    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("rds")

        describe_databases = client.describe_db_instances()
        db_instances = describe_databases["DBInstances"]

        for db_instance in db_instances:
            if "aurora" in db_instance["Engine"]:
                continue

            arn = db_instance.get("DBInstanceIdentifier", "")

            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsRds:{arn}",
                    "region_name": region_name,
                    "arn": db_instance.get("DBInstanceIdentifier", ""),
                    "identifier": db_instance["DBInstanceIdentifier"],
                    "status": db_instance["DBInstanceStatus"],
                    "instance_class": db_instance["DBInstanceClass"],
                    "storage_type": db_instance["StorageType"],
                    "engine": db_instance["Engine"],
                    "engine_version": db_instance["EngineVersion"],
                    # "engine_latest_version": AwsRds.latest_rds_version(
                    #     client, db_instance["Engine"]
                    # ),
                    # "version_is_latest": AwsRds.latest_rds_version(
                    #     client, db_instance["Engine"]
                    # )
                    # == db.engine_version,
                    "deletion_protection": db_instance["DeletionProtection"],
                    "username": (db_instance["MasterUsername"],),
                    "host": (db_instance["Endpoint"]["Address"],),
                    "port": db_instance["Endpoint"]["Port"],
                    "storage_encrypted": db_instance["StorageEncrypted"],
                    "multi_az": db_instance["MultiAZ"],
                    "performance_insights": db_instance["PerformanceInsightsEnabled"],
                    "allocated_storage": db_instance["AllocatedStorage"],
                    # "db_created_at": db_instance["InstanceCreateTime"],
                }
            )

            # db.check_if_instance_class_is_arm()
            # db.check_if_storage_type_is_gp3()
            # db.check_cpu_usage_over_30_days()
            # db.check_db_connections_over_30_days()


#     @staticmethod
#     def latest_rds_version(client, engine):
#         response = client.describe_db_engine_versions(Engine=engine)
#         return response["DBEngineVersions"][-1]["EngineVersion"]

#     def check_if_rds_out_of_date(self, engine):
#         response = self.client.describe_db_instances()
#         for db_instances in response["DBInstances"]:
#             if engine == db_instances["Engine"]:
#                 db_instance_identifier = db_instances["DBInstanceIdentifier"]
#                 engine = db_instances["Engine"]
#                 engine_version = db_instances["EngineVersion"]

#                 if self.latest_rds_version(engine) != engine_version:
#                     print(
#                         f"\tRDS Instance '{db_instance_identifier}' is using '{engine}' engine version {engine_version}. It is out of date\n"
#                     )

#     def check_if_instance_class_is_arm(self):
#         if "g." in self.instance_class:
#             self.instance_class_is_arm = True
#             self.save()

#     def check_if_storage_type_is_gp3(self):
#         if self.storage_type in [
#             "aurora",
#             "gp3",
#         ]:
#             self.storage_type_is_gp3 = True
#             self.save()

#     def check_cpu_usage_over_30_days(self):
#         client = self.aws_account.aws_session(self.aws_region, "cloudwatch")
#         response = client.get_metric_statistics(
#             Namespace="AWS/RDS",
#             MetricName="CPUUtilization",
#             Dimensions=[
#                 {"Name": "DBInstanceIdentifier", "Value": self.identifier},
#             ],
#             StartTime=datetime.utcnow() - timedelta(days=30),
#             EndTime=datetime.utcnow(),
#             Period=3600,
#             Statistics=[
#                 "Average",
#             ],
#         )

#         self.cpu_95th_percentile_over_30_days = pd.Series(
#             [r["Average"] for r in response["Datapoints"]],
#         ).quantile(0.95)
#         self.cpu_99th_percentile_over_30_days = pd.Series(
#             [r["Average"] for r in response["Datapoints"]],
#         ).quantile(0.99)
#         self.save()

#     def check_db_connections_over_30_days(self):
#         client = self.aws_account.aws_session(self.aws_region, "cloudwatch")
#         response = client.get_metric_statistics(
#             Namespace="AWS/RDS",
#             MetricName="DatabaseConnections",
#             Dimensions=[
#                 {"Name": "DBInstanceIdentifier", "Value": self.identifier},
#             ],
#             StartTime=datetime.utcnow() - timedelta(days=30),
#             EndTime=datetime.utcnow(),
#             Period=3600,
#             Statistics=[
#                 "Sum",
#             ],
#         )

#         self.db_connections_95th_percentile_over_30_days = pd.Series(
#             [r["Sum"] for r in response["Datapoints"]],
#         ).quantile(0.95)
#         self.db_connections_99th_percentile_over_30_days = pd.Series(
#             [r["Sum"] for r in response["Datapoints"]],
#         ).quantile(0.99)
#         self.save()


class AwsRdsAurora:
    #     @staticmethod
    #     def latest_rds_version(client, engine):
    #         response = client.describe_db_engine_versions(Engine=engine)
    #         return response["DBEngineVersions"][-1]["EngineVersion"]

    #     def check_if_instance_class_is_arm(self):
    #         if "g." in self.instance_class:
    #             self.instance_class_is_arm = True
    #             self.save()

    #     def check_if_storage_type_is_gp3(self):
    #         if self.storage_type in [
    #             "aurora",
    #             "gp3",
    #         ]:
    #             self.storage_type_is_gp3 = True
    #             self.save()

    #     def check_cpu_usage_over_30_days(self):
    #         client = self.aws_account.aws_session(self.aws_region, "cloudwatch")
    #         response = client.get_metric_statistics(
    #             Namespace="AWS/RDS",
    #             MetricName="CPUUtilization",
    #             Dimensions=[
    #                 {"Name": "DBInstanceIdentifier", "Value": self.identifier},
    #             ],
    #             StartTime=datetime.utcnow() - timedelta(days=30),
    #             EndTime=datetime.utcnow(),
    #             Period=3600,
    #             Statistics=[
    #                 "Average",
    #             ],
    #         )

    #         self.cpu_95th_percentile_over_30_days = pd.Series(
    #             [r["Average"] for r in response["Datapoints"]],
    #         ).quantile(0.95)
    #         self.cpu_99th_percentile_over_30_days = pd.Series(
    #             [r["Average"] for r in response["Datapoints"]],
    #         ).quantile(0.99)
    #         self.save()

    #     def check_db_connections_over_30_days(self):
    #         client = self.aws_account.aws_session(self.aws_region, "cloudwatch")

    #         response = client.get_metric_statistics(
    #             Namespace="AWS/RDS",
    #             MetricName="DatabaseConnections",
    #             Dimensions=[
    #                 {"Name": "DBInstanceIdentifier", "Value": self.identifier},
    #             ],
    #             StartTime=datetime.utcnow() - timedelta(days=30),
    #             EndTime=datetime.utcnow(),
    #             Period=3600,
    #             Statistics=[
    #                 "Sum",
    #             ],
    #         )

    #         self.db_connections_95th_percentile_over_30_days = pd.Series(
    #             [r["Sum"] for r in response["Datapoints"]],
    #         ).quantile(0.95)
    #         self.db_connections_99th_percentile_over_30_days = pd.Series(
    #             [r["Sum"] for r in response["Datapoints"]],
    #         ).quantile(0.99)
    #         self.save()

    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("rds")

        describe_databases = client.describe_db_instances()
        db_instances = describe_databases["DBInstances"]

        for db_instance in db_instances:
            if "aurora" not in db_instance["Engine"]:
                continue

            arn = db_instance.get("DBInstanceIdentifier", "")
            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsRdsAurora:{arn}",
                    "region_name": region_name,
                    "arn": db_instance.get("DBInstanceArn", ""),
                    "identifier": db_instance["DBInstanceIdentifier"],
                    "aws_region": region_name,
                    "status": db_instance["DBInstanceStatus"],
                    "instance_class": db_instance["DBInstanceClass"],
                    "storage_type": db_instance["StorageType"],
                    "engine": db_instance["Engine"],
                    "engine_version": db_instance["EngineVersion"],
                    "engine_latest_version": cls.latest_rds_version(
                        client, db_instance["Engine"]
                    ),
                    # "version_is_latest": cls.latest_rds_version( client, db_instance["Engine"]) == engine_version,
                    "deletion_protection": db_instance["DeletionProtection"],
                    "username": (db_instance["MasterUsername"],),
                    "host": (db_instance["Endpoint"]["Address"],),
                    "port": db_instance["Endpoint"]["Port"],
                    "storage_encrypted": db_instance["StorageEncrypted"],
                    "multi_az": db_instance["MultiAZ"],
                    "performance_insights": db_instance["PerformanceInsightsEnabled"],
                    "allocated_storage": db_instance["AllocatedStorage"],
                    "db_created_at": db_instance["InstanceCreateTime"],
                }
            )
            #  deleted_at = None

            # check_if_instance_class_is_arm()
            # check_if_storage_type_is_gp3()
            # check_cpu_usage_over_30_days()
            # check_db_connections_over_30_days()


class AwsRdsSnapshot:
    #     # @property
    #     # def snapshot_created_at(self):
    #     #     return self.snapshot_create_time.replace(tzinfo=None)

    #     def check_monthly_cost_in_usd(self):
    #         if self.snapshot_type == "automated":
    #             self.monthly_cost_in_usd = None
    #         else:
    #             self.monthly_cost_in_usd = self.allocated_storage * self.SNAPSHOT_COST

    #         self.save()

    #     def check_if_snapshot_younger_than_30_days(self):
    #         if self.snapshot_create_time.replace(tzinfo=None) >= datetime.now().replace(
    #             tzinfo=None
    #         ) - timedelta(days=30):
    #             self.snapshot_younger_than_30_days = True
    #             self.snapshot_older_than_30_days = False
    #             self.save()

    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("rds")

        snapshots = client.describe_db_snapshots()

        for snapshot in snapshots["DBSnapshots"]:
            arn = snapshot.get("DBSnapshotArn", "")
            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsRdsSnapshot:{arn}",
                    "region_name": region_name,
                    "arn": snapshot.get("DBSnapshotArn", ""),
                    "aws_region": region_name,
                    "db_snapshot_identifier": snapshot["DBSnapshotIdentifier"],
                    "db_instance_identifier": snapshot["DBInstanceIdentifier"],
                    "snapshot_create_time": snapshot["SnapshotCreateTime"].replace(
                        tzinfo=None
                    ),
                    "engine": snapshot["Engine"],
                    "allocated_storage": snapshot["AllocatedStorage"],
                    "status": snapshot["Status"],
                    "port": snapshot["Port"],
                    "availability_zone": snapshot["AvailabilityZone"],
                    "vpc_id": snapshot.get("VpcId", None),
                    "instance_create_time": snapshot["InstanceCreateTime"].replace(
                        tzinfo=None
                    ),
                    "master_username": snapshot["MasterUsername"],
                    "engine_version": snapshot["EngineVersion"],
                    "license_model": snapshot["LicenseModel"],
                    "snapshot_type": snapshot["SnapshotType"],
                    "option_group_name": snapshot["OptionGroupName"],
                    "percent_progress": snapshot["PercentProgress"],
                    "storage_type": snapshot["StorageType"],
                    "encrypted": snapshot["Encrypted"],
                    "kms_key_id": snapshot.get("KmsKeyId", None),
                    "db_snapshot_arn": snapshot["DBSnapshotArn"],
                    "iam_database_authentication_enabled": snapshot[
                        "IAMDatabaseAuthenticationEnabled"
                    ],
                    "processor_features": snapshot["ProcessorFeatures"],
                    "dbi_resource_id": snapshot["DbiResourceId"],
                    "tag_list": snapshot["TagList"],
                    "snapshot_target": snapshot["SnapshotTarget"],
                }
            )
            # check_if_snapshot_younger_than_30_days()
            # check_monthly_cost_in_usd()


class AwsRedshift:
    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("redshift")
        clusters = client.describe_clusters()["Clusters"]

        for cluster in clusters:
            arn = cluster["ClusterIdentifier"]
            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsRedshift:{arn}",
                    "region_name": region_name,
                    "cluster_identifier": cluster["ClusterIdentifier"],
                    "node_type": cluster["NodeType"],
                    "cluster_version": cluster["ClusterVersion"],
                    "encrypted": cluster["Encrypted"],
                    "cluster_status": cluster["ClusterStatus"],
                    "number_of_nodes": cluster["NumberOfNodes"],
                    "is_scheduled": cluster.get("SnapshotScheduleState") is not None,
                }
            )

            # aws_redshift.check_cpu_usage_over_30_days()
            # aws_redshift.check_db_connections_over_30_days()


#     def check_cpu_usage_over_30_days(self):
#         client = self.aws_account.aws_session(self.aws_account.aws_region, "cloudwatch")
#         response = client.get_metric_statistics(
#             Namespace="AWS/Redshift",
#             MetricName="CPUUtilization",
#             Dimensions=[
#                 {"Name": "ClusterIdentifier", "Value": self.cluster_identifier},
#             ],
#             StartTime=datetime.utcnow() - timedelta(days=30),
#             EndTime=datetime.utcnow(),
#             Period=3600,
#             Statistics=[
#                 "Average",
#             ],
#         )

#         self.cpu_95th_percentile_over_30_days = pd.Series(
#             [r["Average"] for r in response["Datapoints"]],
#         ).quantile(0.95)
#         self.cpu_99th_percentile_over_30_days = pd.Series(
#             [r["Average"] for r in response["Datapoints"]],
#         ).quantile(0.99)
#         self.save()

#     def check_db_connections_over_30_days(self):
#         client = self.aws_account.aws_session(self.aws_account.aws_region, "cloudwatch")
#         response = client.get_metric_statistics(
#             Namespace="AWS/Redshift",
#             MetricName="DatabaseConnections",
#             Dimensions=[
#                 {"Name": "ClusterIdentifier", "Value": self.cluster_identifier},
#             ],
#             StartTime=datetime.utcnow() - timedelta(days=30),
#             EndTime=datetime.utcnow(),
#             Period=3600,
#             Statistics=[
#                 "Sum",
#             ],
#         )

#         self.db_connections_95th_percentile_over_30_days = pd.Series(
#             [r["Sum"] for r in response["Datapoints"]],
#         ).quantile(0.95)
#         self.db_connections_99th_percentile_over_30_days = pd.Series(
#             [r["Sum"] for r in response["Datapoints"]],
#         ).quantile(0.99)
#         self.save()


class AwsRedshiftSnapshot:
    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("redshift")
        snapshots = client.describe_cluster_snapshots()["Snapshots"]

        for snapshot in snapshots:
            arn = snapshot["SnapshotIdentifier"]
            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsRedshiftSnapshot:{arn}",
                    "region_name": region_name,
                    "snapshot_identifier": snapshot["SnapshotIdentifier"],
                    "cluster_identifier": snapshot["ClusterIdentifier"],
                    "snapshot_create_time": snapshot["SnapshotCreateTime"],
                    "snapshot_type": snapshot["SnapshotType"],
                    "status": snapshot["Status"],
                }
            )


class AwsS3:
    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("s3")

        buckets = client.list_buckets()["Buckets"]

        for bucket in buckets:
            # life_policy = False
            # try:
            #     client.get_bucket_lifecycle(Bucket=bucket["Name"])
            #     life_policy = True
            # except:
            #     life_policy = False

            arn = bucket.get("Name", "")
            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsS3:{arn}",
                    "region_name": region_name,
                    "arn": bucket.get("ARN", ""),
                    "name": bucket["Name"],
                    # "creation_date": bucket["CreationDate"],
                    # "life_policy": life_policy,
                }
            )


class AwsVpcNat:
    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("ec2")
        nat_gateways = client.describe_nat_gateways()["NatGateways"]

        for nat_gateway in nat_gateways:
            arn = nat_gateway.get("NatGatewayId", "")
            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsVpcNat:{arn}",
                    "region_name": region_name,
                    "arn": nat_gateway.get("NatGatewayId", ""),
                    "state": nat_gateway["State"],
                    "connectivity_type": nat_gateway["ConnectivityType"],
                    "subnet_id": nat_gateway["SubnetId"],
                    "vpc_id": nat_gateway["VpcId"],
                    "n_gateways": len(nat_gateway["NatGatewayAddresses"]),
                }
            )


class AwsWaf:
    @classmethod
    def sync(cls, aws_account, region_name):
        client = aws_account.aws_session(region_name).client("waf")
        web_acls = client.list_web_acls()["WebACLs"]

        for web_acl_id in web_acls:
            arn = web_acl_id.get("WebACLId", "")
            ddb_put_item(
                {
                    "customer": aws_account.profile,
                    "service": f"AwsWaf:{arn}",
                    "region_name": region_name,
                    "arn": web_acl_id.get("WebACLId", ""),
                    "name": web_acl_id["Name"],
                    "metric_name": web_acl_id["MetricName"],
                    "n_rules": len(web_acl_id.get("Rules", [])),
                }
            )


class AwsAccount:
    profile = ""

    aws_access_key_id = ""
    aws_secret_access_key = ""
    aws_region_name = ""

    # report_url = models.URLField(blank=True, max_length=1000)
    # report_drive_item_id = models.CharField(blank=True, max_length=1000)

    # created_at = models.DateTimeField(auto_now_add=True)
    # updated_at = models.DateTimeField(auto_now=True)

    def aws_session(self, region_name):
        return boto3.Session(
            profile_name=self.profile,
            region_name=region_name,
        )

    def regions(self):
        return self.aws_session("us-east-1").client("ec2").describe_regions()["Regions"]

    def current_mfa_token(self):
        return pyotp.parse_uri(self.mfa_otp_uri).now()


def report(profile):
    GREEN_FILL = PatternFill(
        start_color="00FF00", end_color="00FF00", fill_type="solid"
    )
    RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    DOCS = [
        {
            "Sheet": "RDS/RDS Aurora",
            "Name": "ARM",
            "Check": "instance_class_is_arm",
            "Description": "ARM Processors are now more performant per cost than x86 instances. Check to see that the instance is using ARM instances",
        },
        {
            "Sheet": "RDS/RDS Aurora",
            "Name": "GP3",
            "Check": "storage_type_is_gp3",
            "Description": "Storage type of GP3 has the best performance compared to other storage types compared to the cost.",
        },
        {
            "Sheet": "RDS/RDS Aurora/EC2/Redshift",
            "Name": "CPU 95th Percentile",
            "Check": "cpu_95th_percentile_over_30_days",
            "Description": "The average CPU usage over the past 30 days. The best approach is to keep the CPU at about 40% usage or more on average otherwise the instance may be overprovisioned.",
        },
        {
            "Sheet": "RDS/RDS Aurora/EC2/Redshift",
            "Name": "CPU 99th Percentile",
            "Check": "cpu_99th_percentile_over_30_days",
            "Description": "The average CPU usage over the past 30 days. The best approach is to keep the CPU at about 40% usage or more on average otherwise the instance may be overprovisioned.",
        },
        {
            "Sheet": "RDS/RDS Aurora/Redshift",
            "Name": "Database Connections 95th Percentile",
            "Check": "db_connections_95th_percentile_over_30_days",
            "Description": "The total database connections over the past 30 days.",
        },
        {
            "Sheet": "RDS/RDS Aurora/Redshift",
            "Name": "Database Connections 99th Percentile",
            "Check": "db_connections_99th_percentile_over_30_days",
            "Description": "The total database connections over the past 30 days.",
        },
        {
            "Sheet": "RDS/RDS Aurora",
            "Name": "Encryption",
            "Check": "storage_encrypted",
            "Description": "Specifies whether the DB cluster is encrypted.",
        },
        {
            "Sheet": "RDS/RDS Aurora",
            "Name": "Deletion Protection",
            "Check": "deletion_protection",
            "Description": "Indicates if the DB cluster has deletion protection enabled. The database can’t be deleted when deletion protection is enabled.",
        },
        {
            "Sheet": "RDS/RDS Aurora",
            "Name": "Allocated Storage",
            "Check": "allocated_storage",
            "Description": "Specifies the allocated storage size in gibibytes (GiB). For Aurora, AllocatedStorage always returns 1, because Aurora DB cluster storage size isn’t fixed.",
        },
        {
            "Sheet": "RDS/RDS Aurora",
            "Name": "Performance Insights",
            "Check": "performance_insights",
            "Description": "Whether Performance Insights is enabled for the DB cluster. This setting is only for non-Aurora Multi-AZ DB clusters.",
        },
        {
            "Sheet": "RDS/RDS Aurora",
            "Name": "Latest Version",
            "Check": "version_is_latest",
            "Description": "Whether the engine version is the latest version of the engine.",
        },
        {
            "Sheet": "RDS Snapshots",
            "Name": "Age",
            "Check": "snapshot_younger_than_30_days",
            "Description": "If a snapshot is older than 30 days you likely don't need it. We look to clean up and remove old snapshots.",
        },
        {
            "Sheet": "RDS Snapshots",
            "Name": "Cost",
            "Check": "monthly_cost_in_usd",
            "Description": "The amount that the RDS snapshots cost per month.",
        },
        {
            "Sheet": "EKS Nodegroups",
            "Name": "Upgrade",
            "Check": "is_upgraded",
            "Description": "whether the nodegroup is upgraded to match the cluster environment version",
        },
        {
            "Sheet": "EKS",
            "Name": "Latest Version",
            "Check": "version_is_latest",
            "Description": "whether the version of the cluster is the latest",
        },
        {
            "Sheet": "EKS Nodegroups",
            "Name": "Spot Instance",
            "Check": "is_spot",
            "Description": "If the nodegroup is a spot instance or not",
        },
        {
            "Sheet": "EKS Nodegroups",
            "Name": "Running BottleRocket",
            "Check": "is_running_bottlerocket",
            "Description": "If the nodegroup ami is running linux bottlerocket.",
        },
        {
            "Sheet": "EC2 ELB",
            "Name": "Number of Instances",
            "Check": "n_instances",
            "Description": "The number of instances that balance the load. If zero and classic ELB, then the ELB is idle",
        },
        {
            "Sheet": "EC2 ELB",
            "Name": "Classic ELB",
            "Check": "not_classic_elb",
            "Description": "Whether the load balancer is classic, i.e. v1 rather than v2.",
        },
        {
            "Sheet": "ECR",
            "Name": "Life Cycle Policy",
            "Check": "lifecycle_policy",
            "Description": "Whether the ECR has a lifecycle policy or not.",
        },
        {
            "Sheet": "VPC NAT",
            "Name": "Number of Gateways",
            "Check": "n_gateways",
            "Description": "number of NAT gateways used in the VPC. If zero, then the gateway is idle.",
        },
        {
            "Sheet": "Dynamo DB",
            "Name": "Deletion Protection",
            "Check": "deletion_protection",
            "Description": "Whether deletion protection is enabled on the table.",
        },
        {
            "Sheet": "EC2 EBS Volume",
            "Name": "Volume Encryption",
            "Check": "encrypted",
            "Description": "Whether the EBS volume is encrypted.",
        },
        {
            "Sheet": "EC2 EBS Volume",
            "Name": "Volume Type",
            "Check": "volume_type",
            "Description": "Type of the volume, e.g. gp2 or gp3.",
        },
        {
            "Sheet": "EC2 EBS Volume",
            "Name": "Connected",
            "Check": "is_attached",
            "Description": "Whether there are any ec2 instances that are attached to the volume.",
        },
        {
            "Sheet": "EC2 EBS Snapshot",
            "Name": "Age",
            "Check": "younger_than_30_days",
            "Description": "If a snapshot is older than 30 days you likely don't need it. We look to clean up and remove old snapshots.",
        },
        {
            "Sheet": "Elasticache",
            "Name": "Version",
            "Check": "is_latest_version",
            "Description": "Whether the version of the cache engine is the latest.",
        },
        {
            "Sheet": "Elasticache",
            "Name": "Transit Encryption",
            "Check": "transit_encryption",
            "Description": "Whether transit encryption is enabled.",
        },
        {
            "Sheet": "Elasticache",
            "Name": "ARM",
            "Check": "is_arm",
            "Description": "Whether the cache node type is using an ARM processor.",
        },
        {
            "Sheet": "Elasticache",
            "Name": "At Rest Encryption",
            "Check": "at_rest_encryption",
            "Description": "Whether at rest encryption is enabled.",
        },
        {
            "Sheet": "CloudWatch Log Group",
            "Name": "Retention In Days",
            "Check": "retention_in_days",
            "Description": "The number of days to retain the log events in the specified log group. null value implies no expiry date, i.e. no life cycle",
        },
        {
            "Sheet": "CloudWatch Log Group",
            "Name": "Lifecycle",
            "Check": "has_lifecycle",
            "Description": "whether the log group has an expiry date",
        },
        {
            "Sheet": "Redshift",
            "Name": "Scheduled",
            "Check": "is_scheduled",
            "Description": "whether the redshift cluster has a schedule or not",
        },
        {
            "Sheet": "Lambda",
            "Name": "ARM",
            "Check": "is_arm",
            "Description": "whether the lambda architecture is using ARM processor or not",
        },
        {
            "Sheet": "Lambda",
            "Name": "Valid Runtime",
            "Check": "is_valid_runtime",
            "Description": "whether the runtime of the lambda function is valid or deprecated",
        },
    ]

    model_to_fields = [
        (
            "AwsRds",
            [
                "arn",
                "aws_region",
                "identifier",
                "engine",
                "instance_class",
                "instance_class_is_arm",
                "storage_type",
                "storage_type_is_gp3",
                "cpu_95th_percentile_over_30_days",
                "cpu_99th_percentile_over_30_days",
                "db_connections_95th_percentile_over_30_days",
                "db_connections_99th_percentile_over_30_days",
                "storage_encrypted",
                "deletion_protection",
                "performance_insights",
                "allocated_storage",
                "multi_az",
                "version_is_latest",
                "db_created_at",
            ],
        ),
        (
            "AwsRdsAurora",
            [
                "arn",
                "aws_region",
                "identifier",
                "engine",
                "instance_class",
                "instance_class_is_arm",
                "storage_type",
                "storage_type_is_gp3",
                "cpu_95th_percentile_over_30_days",
                "cpu_99th_percentile_over_30_days",
                "db_connections_95th_percentile_over_30_days",
                "db_connections_99th_percentile_over_30_days",
                "storage_encrypted",
                "deletion_protection",
                "performance_insights",
                "allocated_storage",
                "multi_az",
                "version_is_latest",
            ],
        ),
        (
            "AwsRdsSnapshot",
            [
                "aws_region",
                "db_snapshot_identifier",
                "snapshot_type",
                "allocated_storage",
                "monthly_cost_in_usd",
                "snapshot_younger_than_30_days",
                "snapshot_create_time",
            ],
        ),
        (
            "AwsCloudwatchLogGroup",
            [
                "arn",
                "log_group_name",
                "data_protection_status",
                "stored_gbs",
                "retention_in_days",
                "has_lifecycle",
            ],
        ),
        (
            "AwsEks",
            [
                "arn",
                "environment",
                "aws_region",
                "cluster_version",
                "latest_cluster_version",
                "version_is_latest",
                "spot_instance_percentage",
                "node_versions_upgraded",
                "karpenter_percentage",
            ],
        ),
        (
            "AwsEksNodegroup",
            [
                "arn",
                "aws_eks__environment",
                "name",
                "version",
                "release_version",
                "status",
                "scaling_config",
                "instance_types",
                "is_upgraded",
                "is_spot",
                "is_running_bottlerocket",
            ],
        ),
        (
            "AwsEc2",
            [
                "arn",
                "aws_instance_id",
                "aws_instance_type",
                "aws_platform_details",
                "cpu_95th_percentile_over_30_days",
                "cpu_99th_percentile_over_30_days",
                "is_spot_instance",
                "is_eks_node",
            ],
        ),
        ("AwsEc2Elb", ["arn", "name", "type", "n_instances", "not_classic_elb"]),
        (
            "AwsEcr",
            [
                "arn",
                "name",
                "mutability",
                "lifecycle_policy",
                "total_storage_used_in_gb",
            ],
        ),
        (
            "AwsVpcNat",
            ["arn", "gateway_id", "state", "connectivity_type", "n_gateways"],
        ),
        (
            "AwsDynamoDB",
            ["arn", "table_id", "table_name", "table_status", "deletion_protection"],
        ),
        (
            "AwsEc2EbsVolume",
            [
                "arn",
                "volume_id",
                "volume_type",
                "encrypted",
                "is_attached",
                "is_storage_type_gp3",
                "allocation_size",
            ],
        ),
        (
            "AwsEc2EbsSnapshot",
            [
                "arn",
                "snapshot_id",
                "encrypted",
                "volume_size",
                "younger_than_30_days",
            ],
        ),
        (
            "AwsElasticache",
            [
                "arn",
                "cache_cluster_id",
                "engine",
                "engine_version",
                "latest_engine_version",
                "is_latest_version",
                "transit_encryption",
                "at_rest_encryption",
                "is_arm",
            ],
        ),
        (
            "AwsRedshift",
            [
                "cluster_identifier",
                "encrypted",
                "is_scheduled",
                "number_of_nodes",
                "cpu_95th_percentile_over_30_days",
                "cpu_99th_percentile_over_30_days",
                "db_connections_95th_percentile_over_30_days",
                "db_connections_99th_percentile_over_30_days",
            ],
        ),
        (
            "AwsRedshiftSnapshot",
            [
                "snapshot_identifier",
                "cluster_identifier",
                "snapshot_type",
            ],
        ),
        (
            "AwsCloudformationStack",
            (
                "arn",
                "stack_id",
                "stack_name",
                "stack_status",
                "role_arn",
                "disable_rollback",
                "termination_protection",
            ),
        ),
        (
            "AwsCloudFrontDistribution",
            (
                "domain_name",
                "distribution_id",
                "status",
                "enabled",
                "staging",
                "is_ipv6_enabled",
            ),
        ),
        (
            "AwsLambda",
            (
                "function_name",
                "role",
                "run_time",
                "handler",
                "version",
                "is_arm",
                "is_valid_runtime",
            ),
        ),
        ("AwsWaf", ("acl_name", "acl_metric_name", "n_rules")),
    ]

    def clean_up_cols(ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)

                    # Boolean values
                    if cell.value == True:
                        cell.fill = GREEN_FILL
                    elif cell.value == False:
                        cell.fill = RED_FILL
                except:
                    pass
                adjusted_width = max_length + 20
                ws.column_dimensions[column].width = adjusted_width

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        print(tmp.name)
        with pd.ExcelWriter(tmp.name) as writer:
            df = pd.DataFrame.from_records(DOCS)
            df.to_excel(writer, sheet_name="Docs")
            clean_up_cols(writer.sheets["Docs"])

            for model, fields in model_to_fields:
                try:
                    from boto3.dynamodb.conditions import Key, Attr

                    df = pd.DataFrame(
                        ddb_table().query(
                            KeyConditionExpression=Key("customer").eq(profile)
                            & Key("service").begins_with(f"{model}:")
                        )["Items"]
                    )

                    if len(df) == 0:
                        continue

                    df.to_excel(writer, sheet_name=model)
                    clean_up_cols(writer.sheets[model])
                except KeyError:
                    pass

        # now = datetime.datetime.now().strftime("%Y-%m-%d-%H%M")
        # sharepoint_file = ms_graph_upload(
        #     local_file_name=tmp.name,
        #     remote_file_name=f"/Product/omyac/{now}-{aws_account.profile}.xlsx",
        # )
        # aws_account.omyac_report_drive_item_id = sharepoint_file["ms_drive_item_id"]
        # aws_account.omyac_report_url = ms_graph_file_share(
        #     aws_account.omyac_report_drive_item_id, "edit", "anonymous"
        # )

        # aws_account.save()

        # # check AwsEksNodegroup Version
        # for cluster in AwsEks.objects.all():
        #     nodegroups = AwsEksNodegroup.objects.filter(
        #         aws_eks=cluster, aws_account=aws_account
        #     )
        #     if nodegroups.exists():
        #         # Get the version of the first nodegroup for this EKS instance
        #         first_nodegroup_version = nodegroups.first().version

        #         # Check if all nodegroups for this EKS instance have the same version
        #         if all(
        #             nodegroup.version == first_nodegroup_version
        #             for nodegroup in nodegroups
        #         ):
        #             print(
        #                 f"EKS instance '{cluster.environment}' has all nodegroups with version {first_nodegroup_version}"
        #             )
        #             cluster.node_versions_upgraded = True

        #         else:
        #             print(
        #                 f"EKS instance '{cluster.environment}' has nodegroups with different versions"
        #             )
        #             cluster.node_versions_upgraded = False
        #         cluster.save()

        # if lead.slack_channel:
        #     lead.notify_on_slack(
        #         f"OMYAC Report: {aws_account.profile} - {aws_account.omyac_report_url}"
        #     )


def sync_region(c, region_name):
    AwsCloudformationStack.sync(c, region_name)
    AwsEks.sync(c, region_name)
    AwsDynamoDB.sync(c, region_name)
    AwsCloudFrontDistribution.sync(c, region_name)
    AwsCloudwatchLogGroup.sync(c, region_name)
    AwsEc2.sync(c, region_name)
    AwsEc2Ami.sync(c, region_name)
    AwsEc2AutoScalingGroup.sync(c, region_name)
    AwsEc2EbsSnapshot.sync(c, region_name)
    AwsEc2EbsVolume.sync(c, region_name)
    AwsEc2Elb.sync(c, region_name)
    AwsEcr.sync(c, region_name)
    AwsEksNodegroup.sync(c, region_name)
    AwsElasticache.sync(c, region_name)
    AwsElasticacheSnapshot.sync(c, region_name)
    AwsLambda.sync(c, region_name)
    AwsRds.sync(c, region_name)
    AwsRdsAurora.sync(c, region_name)
    AwsRdsSnapshot.sync(c, region_name)
    AwsRedshift.sync(c, region_name)
    AwsRedshiftSnapshot.sync(c, region_name)
    AwsS3.sync(c, region_name)
    AwsVpcNat.sync(c, region_name)
    AwsWaf.sync(c, region_name)


from multiprocessing import Pool

if __name__ == "__main__":
    c = AwsAccount()
    c.profile = "opszero"

    # with Pool(10) as pool:
    #     pool.starmap(
    #         sync_region,
    #         [
    #             (
    #                 c,
    #                 region["RegionName"],
    #             )
    #             for region in c.regions()
    #         ],
    #     )

    report(c.profile)
